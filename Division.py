# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
import csv
import datetime
import re
import os.path
import glob
import Library
from shutil import copyfile
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import formatting, styles
import configparser

def MainDivision():
    try:
        print('************ BEGIN ************')
        print()
        print('Step 1: Processing Template and Raw Data')
        try:
            try:
                config = configparser.ConfigParser()
                config.read('Settings.ini')
                rawdata_path=config.get('source', 'rawdata')
                if rawdata_path=="": rawdata_path=None
            except: rawdata_path=None

            #1.import template
            wb_file_name = 'Report\AG_Performance_Template.xlsx'
            wb_file = Library.getXlsxFile(wb_file_name, [])
            print('  -Loading file: %s ' % wb_file, end="")
            wb = load_workbook(filename = wb_file)
            wb_sheet = wb["Performance"]
            ReportDate = wb_sheet.cell(column=1, row=1).value
            ReportDateSimpleArr = str(ReportDate).split(' ',1)[0].split('-',2)
            ReportDateSimple = ReportDateSimpleArr[1]+ReportDateSimpleArr[2]
            ReportDateStr = ReportDateSimpleArr[0]+"/"+ReportDateSimpleArr[1]+"/"+ReportDateSimpleArr[2]
            
            for checksheet in wb.sheetnames:
                if checksheet!="分時表":
                    wb.remove(wb[checksheet])
            wb_sheet = wb["分時表"]
            print(' => Completed')

            #2.import data source
            #Skill1
            print('  -Loading file:', end="")
            table_skill1_file=None
            if rawdata_path!=None:
                try:
                    table_skill1_file_name = rawdata_path + '\**\*間隔skill總結*.xls'
                    table_skill1_file = Library.getCsvFile(table_loginout_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_skill1_file=None
            if table_skill1_file==None:
                table_skill1_file_name = 'Report\RAWDATA\**\*間隔skill總結*.xls'
                table_skill1_file = Library.getCsvFile(table_skill1_file_name, [[1,1, ReportDateStr], [1,2, 'OPPO SKILL 1']])

            if table_skill1_file!=None:
                table_skill1 = csv.reader(open(table_skill1_file, 'r'), delimiter='\t')   
                table_skill1_list = list(table_skill1) if table_skill1!=None else None
                table_skill1_rows = sum(1 for row in table_skill1) if table_skill1_list!=None else 0
                print(' %s => Completed' % table_skill1_file)
            else:
                print(' Finding "%s\" => Failed' % table_skill1_file_name)

            #Skill2
            print('  -Loading file:', end="")
            table_skill2_file=None
            if rawdata_path!=None:
                try:
                    table_skill2_file_name = rawdata_path + '\**\*間隔skill總結*.xls'
                    table_skill2_file = Library.getCsvFile(table_skill2_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_skill2_file=None
            if table_skill2_file==None:
                table_skill2_file_name = 'Report\RAWDATA\**\*間隔skill總結*.xls'
                table_skill2_file = Library.getCsvFile(table_skill2_file_name, [[1,1, ReportDateStr], [1,2, 'OPPO SKILL 2']])

            if table_skill2_file!=None:
                table_skill2 = csv.reader(open(table_skill2_file, 'r'), delimiter='\t')   
                table_skill2_list = list(table_skill2) if table_skill2!=None else None
                table_skill2_rows = sum(1 for row in table_skill2) if table_skill2_list!=None else 0
                print(' %s => Completed' % table_skill2_file)
            else:
                print(' Finding "%s\" => Failed' % table_skill2_file_name)

            #Skill3
            print('  -Loading file:', end="")
            table_skill3_file=None
            if rawdata_path!=None:
                try:
                    table_skill3_file_name = rawdata_path + '\**\*間隔skill總結*.xls'
                    table_skill3_file = Library.getCsvFile(table_skill3_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_skill3_file=None
            if table_skill3_file==None:
                table_skill3_file_name = 'Report\RAWDATA\**\*間隔skill總結*.xls'
                table_skill3_file = Library.getCsvFile(table_skill3_file_name, [[1,1, ReportDateStr], [1,2, 'OPPO SKILL 3']])
                
            if table_skill3_file!=None:
                table_skill3 = csv.reader(open(table_skill3_file, 'r'), delimiter='\t')   
                table_skill3_list = list(table_skill3) if table_skill3!=None else None
                table_skill3_rows = sum(1 for row in table_skill3) if table_skill3_list!=None else 0
                print(' %s => Completed' % table_skill3_file)
            else:
                print(' Finding "%s\" => Failed' % table_skill3_file_name)

            #Skill4
            print('  -Loading file:', end="")
            table_skill4_file=None
            if rawdata_path!=None:
                try:
                    table_skill4_file_name = rawdata_path + '\**\*間隔skill總結*.xls'
                    table_skill4_file = Library.getCsvFile(table_skill4_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_skill4_file=None
            if table_skill4_file==None:
                table_skill4_file_name = 'Report\RAWDATA\**\*間隔skill總結*.xls'
                table_skill4_file = Library.getCsvFile(table_skill4_file_name, [[1,1, ReportDateStr], [1,2, 'OPPO SKILL 4']])

            if table_skill4_file!=None:
                table_skill4 = csv.reader(open(table_skill4_file, 'r'), delimiter='\t')   
                table_skill4_list = list(table_skill4) if table_skill4!=None else None
                table_skill4_rows = sum(1 for row in table_skill4) if table_skill4_list!=None else 0
                print(' %s => Completed' % table_skill4_file)
            else:
                print(' Finding "%s\" => Failed' % table_skill4_file_name)
                
            #VM
            print('  -Loading file:', end="")
            table_VM_file=None
            if rawdata_path!=None:
                try:
                    table_VM_file_name = rawdata_path + '\**\*間隔skill總結*.xls'
                    table_VM_file = Library.getCsvFile(table_skill4_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_VM_file=None
            if table_VM_file==None:
                table_VM_file_name = 'Report\RAWDATA\**\*間隔skill總結*.xls'
                table_VM_file = Library.getCsvFile(table_VM_file_name, [[1,0, ReportDateStr], [1,2, 'OPPO to IVR 2040000']])

            if table_VM_file!=None:
                table_VM = csv.reader(open(table_VM_file, 'r'), delimiter='\t')   
                table_VM_list = list(table_VM) if table_VM!=None else None
                table_VM_rows = sum(1 for row in table_VM) if table_VM_list!=None else 0
                print(' %s => Completed' % table_VM_file)
            else:
                print(' Finding "%s\" => Failed' % table_VM_file_name)

            

            #for skill1 in table_skill1_list:
            #    print(Library.convertTimeDivisionFull(skill1[0], skill1[1]))


            #3.Processing Report
            print()
            print('Step 2: Processing report calculation')

            wb_sheet.cell(column=1, row=2).value=ReportDateStr
            Totaltalkingsec=0
            TotalACD=0
            TotalACW=0
            TotalHoldsec=0
            TotalHold=0
            TotalAcw=0
            TotalPickupSec=0
            for rows in range(3, wb_sheet.max_row+1):
                FullDataTime = ReportDateStr + ' ' + wb_sheet.cell(column=2, row=rows).value #日期/時間
                print('  -Loading Time: %s' % FullDataTime, end="")
                wb_sheet.cell(column=1, row=rows).value=FullDataTime
                wb_sheet.cell(column=4, row=rows).value=0 #ACD通話
                wb_sheet.cell(column=5, row=rows).value=0 #掛斷通話
                wb_sheet.cell(column=6, row=rows).value=0 #Voice Mail
                wb_sheet.cell(column=10, row=rows).value=0 #平均位置值班
                talkingsec=0
                holdsec=0 #等候時間
                hold=0 #等候通話
                acwsec=0 #ACW
                pickupsec=0 #接聽時間
                #Skill1            
                for skill1 in table_skill1_list:
                    if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill1[0], skill1[1]):
                        wb_sheet.cell(column=4, row=rows).value+=int(skill1[4]) #ACD通話
                        wb_sheet.cell(column=5, row=rows).value+=int(skill1[7]) #掛斷通話
                        wb_sheet.cell(column=10, row=rows).value+=int(skill1[17]) #平均位置值班
                        holdsec+=int(skill1[20]) #等候時間
                        hold+=int(skill1[21]) #等候通話
                        TotalACD+=int(skill1[4])
                        talkingsec+=float(skill1[4])*float(skill1[5]) #ACD通話*平均ACD時間
                        acwsec+=float(skill1[4])*float(skill1[6]) #ACD通話*平均ACW時間
                        pickupsec+=float(skill1[4])*float(skill1[2]) #ACD通話*平均速度接聽
                        break            
                #Skill2
                for skill2 in table_skill2_list:                
                    if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill2[0], skill2[1]):
                        wb_sheet.cell(column=4, row=rows).value+=int(skill2[4]) #ACD通話
                        wb_sheet.cell(column=5, row=rows).value+=int(skill2[7]) #掛斷通話
                        holdsec+=int(skill2[20]) #等候時間
                        hold+=int(skill2[21]) #等候通話
                        TotalACD+=int(skill2[4])
                        talkingsec+=float(skill2[4])*float(skill2[5]) #ACD通話*平均ACD時間
                        acwsec+=float(skill2[4])*float(skill2[6]) #ACD通話*平均ACW時間
                        pickupsec+=float(skill2[4])*float(skill2[2]) #ACD通話*平均速度接聽
                        break
                #Skill3
                for skill3 in table_skill3_list:
                    if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill3[0], skill3[1]):
                        wb_sheet.cell(column=4, row=rows).value+=int(skill3[4]) #ACD通話
                        wb_sheet.cell(column=5, row=rows).value+=int(skill3[7]) #掛斷通話
                        holdsec+=int(skill3[20]) #等候時間
                        hold+=int(skill3[21]) #等候通話
                        TotalACD+=int(skill3[4])
                        talkingsec+=float(skill3[4])*float(skill3[5]) #ACD通話*平均ACD時間
                        acwsec+=float(skill3[4])*float(skill3[6]) #ACD通話*平均ACW時間
                        pickupsec+=float(skill3[4])*float(skill3[2]) #ACD通話*平均速度接聽
                        break
                #Skill4
                for skill4 in table_skill4_list:
                    if wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFull(skill4[0], skill4[1]):
                        wb_sheet.cell(column=4, row=rows).value+=int(skill4[4]) #ACD通話
                        wb_sheet.cell(column=5, row=rows).value+=int(skill4[7]) #掛斷通話
                        holdsec+=int(skill4[20]) #等候時間
                        hold+=int(skill4[21]) #等候通話
                        TotalACD+=int(skill4[4])
                        talkingsec+=float(skill4[4])*float(skill4[5]) #ACD通話*平均ACD時間
                        acwsec+=float(skill4[4])*float(skill4[6]) #ACD通話*平均ACW時間
                        pickupsec+=float(skill4[4])*float(skill4[2]) #ACD通話*平均速度接聽
                        break
                #VM
                for vm in table_VM_list:
                    if len(vm)>2 and wb_sheet.cell(column=2, row=rows).value==Library.convertTimeDivisionFullB(vm[0], vm[2]): #Csv前幾筆只有兩個欄位
                        wb_sheet.cell(column=6, row=rows).value+=int(vm[14]) #Voice Mail
                        break
                
                #Skill1--
                Totaltalkingsec+=talkingsec
                TotalHoldsec+=holdsec
                TotalHold+=hold
                TotalAcw+=acwsec
                TotalPickupSec+=pickupsec
                try: wb_sheet.cell(column=12, row=rows).value=(talkingsec/float(wb_sheet.cell(column=4, row=rows).value))/3600/24 #ACD通話*平均ACD時間
                except: wb_sheet.cell(column=12, row=rows).value="--"
                try: wb_sheet.cell(column=13, row=rows).value=(holdsec/hold)/3600/24 #ACD通話*平均ACD時間
                except: wb_sheet.cell(column=13, row=rows).value="--"
                try: wb_sheet.cell(column=14, row=rows).value=(holdsec/wb_sheet.cell(column=4, row=rows).value)/3600/24 #ACW通話*平均ACW時間
                except: wb_sheet.cell(column=14, row=rows).value="--"
                try: wb_sheet.cell(column=15, row=rows).value=(holdsec/wb_sheet.cell(column=4, row=rows).value)/3600/24 #ACD通話*平均速度接聽
                except: wb_sheet.cell(column=15, row=rows).value="--"
        
                print(' => Completed')
            
            wb_sheet.cell(column=12, row=2).value=(Totaltalkingsec/TotalACD)/3600/24 #ACD通話/ACD
            wb_sheet.cell(column=13, row=2).value=(TotalHoldsec/TotalHold)/3600/24 #等候時間/等候通話
            wb_sheet.cell(column=14, row=2).value=(TotalAcw/TotalACD)/3600/24 #ACW通話/ACD
            wb_sheet.cell(column=15, row=2).value=(TotalPickupSec/TotalACD)/3600/24 #速度接聽/ACD

            print()
            print('Step 3: Generating Report')
            #Generate        
            wb_sheet.title = ReportDateSimple
            Performance_FilePathName = "Report\分時表%s.xlsx" % ReportDateSimple

            print('  -Creating Report to the %s'  % Performance_FilePathName, end="")

            wb.save(Performance_FilePathName)
            
            wb.close()
            print(' => Completed')    
        #except Exception as e:
        #    print(e)
        except:
            print('  -Loading file: => Failed ')
    except Exception as e:
        print(e)
    #except:
    #   print("Error!! Close all excel files and try again.")

    finally:
        print()
        print('************ END ************')