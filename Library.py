# -*- coding: UTF-8 -*-
from openpyxl import load_workbook
import csv
import datetime
import re
import os.path
import glob
from shutil import copyfile

def getRow(table, loginID):
    myloginID = str(loginID)
    if len(myloginID)==7:
        myloginID = 'A'+myloginID
    for i in table:
        if i[0]==myloginID:
        #if i[0][-7:]==str(myloginID):
            return i
    return None

def formatDateToStr(importDate):
    return str(importDate).split(' ',1)[0].replace("-", "/")

def getMailCount(table, loginName, reportDate):
    try:
        myCount = 0
        #reportDate.replace(hour=0, minute=0, second=0, microsecond=0)
        if table.max_row>1 and loginName!=None:
            for rows in range(2, table.max_row+1):            
                myLoginName = table.cell(column=5, row=rows).value
                if myLoginName!=None:
                    myLoginName = re.sub(r'[\x00-\x7f]',r' ', myLoginName).strip()
                    myMailDate = formatDateToStr(table.cell(column=4, row=rows).value)
                    #myMailDate = myMailDate.replace(hour=0, minute=0, second=0, microsecond=0)
                    if table.cell(column=2, row=rows).value=='Closed' and loginName==myLoginName and reportDate==myMailDate: 
                        myCount += 1
        else:
            return None   
        
        return myCount if myCount>0 else None        
    except:
        return None

def getCtsCount(table, loginName, reportDate, serviceWay):
    try:
        myCount = 0
        #myReportDate = formatDateToStr(reportDate)
        if table.max_row>1 and loginName!=None and serviceWay!=None and reportDate!=None:
            for rows in range(2, table.max_row+1):            
                myLoginName = table.cell(column=20, row=rows).value.replace('　', '').strip()
                #print('Cts Name=%s' % myLoginName)
                if myLoginName!=None:
                    LoginName = re.sub(r'[\x00-\x7f]',r' ', loginName).strip()
                    myMailDate = formatDateToStr(table.cell(column=3, row=rows).value)
                    #print('LoginName=%s, myLoginName=%s, reportDate=%s, myMailDate=%s, serviceWay1=%s, serviceWay2=%s' % (LoginName, myLoginName, myReportDate, myMailDate, table.cell(column=2, row=rows).value, serviceWay))
                    if table.cell(column=2, row=rows).value==serviceWay and LoginName==myLoginName and reportDate==myMailDate: 
                        myCount += 1
        else:
            return None
            
        return myCount if myCount>0 else None
    except:
        print(' => CTS Data Abnormal', end="")
        return None

def getSec(iMinSec):
    try:
        minsec = iMinSec[:-2].split(':',1)
        return int(minsec[1])+(int(minsec[0])*60)+(12*60 if iMinSec[-2:]=='下午' else 0)
    except:
        print(' => Login/out Time Abnormal', end="")
        return 0

def convertTimeDivision(segment):
    try:
        minstr = str(segment)[-2:]
        hourstr = str(segment)[:-2]
        if hourstr == '': hourstr = '00'
        if minstr == '60':
            hourstr = str(int(hourstr)+1)
            minstr = '00'
        return hourstr.zfill(2) + ":" + minstr.zfill(2)
    except:
        return '00:00-00:00'

def convertTimeDivisionFull(start, end):
    return convertTimeDivision(start) + "-" + convertTimeDivision(end)

def convertTimeDivisionFullB(start, end):
    try:
        starthour = int(start.split(':',1)[0])
        startmin  = int(start.split(':',1)[1])
        endhour   = int(end.split(':',1)[0])
        endmin    = int(end.split(':',1)[1][:-2])
        if end.split(':',1)[1][-2:]=='上午' and endhour==12: endhour = starthour = 0
        if end.split(':',1)[1][-2:]=='下午' and starthour<12 and not (starthour==11 and startmin==30): starthour += 12
        if end.split(':',1)[1][-2:]=='下午' and endhour<12: endhour += 12
        if end=='12:00上午': 
            starthour = 11
            endhour = 24
        return str(starthour).zfill(2) + ":" + str(startmin).zfill(2) + "-" + str(endhour).zfill(2) + ":" + str(endmin).zfill(2)
    except:
        return '00:00-00:00'

def copyWorksheet(source, target):
    for rows in range(1, source.max_row+100):
        for cols in range(1, source.max_column+1):
            target.cell(column=cols, row=rows).value = source.cell(column=cols, row=rows).value            
            #if target.cell(column=cols, row=rows).value==0:
            #    target.cell(column=cols, row=rows).style.font.color.index = Color.Gray
    return target    

def getCsvFile(file, conditions):
    for filename in glob.iglob(file, recursive=True):
        try:
            myTable = csv.reader(open(filename, 'r'), delimiter='\t')
            myTable_list = list(myTable)
            if len(conditions)==0: return filename
            myFlag=0
            for conditon in conditions:                
                if myTable_list[conditon[1]][conditon[0]]==conditon[2]: myFlag=1
                else: 
                    myFlag=0
                    break
            if myFlag: return filename
        except:
            print(' => Data Exception')
    return None

def getXlsxFile(file, conditions):
    for filename in glob.iglob(file, recursive=True):
        try:
            #print(filename)
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]
            if len(conditions)==0: return filename
            for rows in range(2, table_sheet.max_row+1):
                myFlag = 0                
                for myCondition in conditions:
                    if myCondition[0] == 'DateStr' and myCondition[2] == str(table_sheet.cell(column=myCondition[1], row=rows).value).split(' ',1)[0].replace("-", "/"):
                        myFlag = 1
                    elif myCondition[0] == 'Str' and myCondition[2] == table_sheet.cell(column=myCondition[1], row=rows).value:
                        myFlag = 1
                    else:
                        myFlag = 0
                if myFlag == 1:
                    table.close()
                    return filename                
            table.close()            
        except:
            print(' => Data Exception')
    return None

def xxxxgetCsvFileName(file, myDate):
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        ##print(' %s' % filename, end="")
        try:
            myTable = csv.reader(open(filename, 'r'), delimiter='\t')
            myTable_list = list(myTable)
            if myTable_list[1][1]==myDate:
                ##print(' => Completed')
                return filename
        except:
            print(' => Data Exception')
    return None

def xxxxgetMailFileName(file, myDate):  
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        #print(' %s' % filename, end="")
        try:
            #print(filename)
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]       
            for rows in range(2, table_sheet.max_row+1):  
                myFileDate = formatDateToStr(table_sheet.cell(column=4, row=rows).value)
                if table_sheet.cell(column=2, row=rows).value=='Closed' and myDate==myFileDate:
                    #print(' => Completed')
                    return filename
        except:
            print(' => Data Exception')
    return None

def xxxxgetCtsFileName(file, myDate):
    for filename in glob.iglob('Report\RAWDATA\**\%s' % file, recursive=True):
        #print(' %s' % filename, end="")
        try:
            table = load_workbook(filename = filename) 
            table_sheet = table[table.sheetnames[0]]
            for rows in range(2, table_sheet.max_row+1):  
                myFileDate = formatDateToStrtable_sheet.cell(column=3, row=rows).value
                if myDate==myFileDate:                    
                    return filename
        except:
            print(' => Data Exception')
        #except Exception as e:
        #    print(e)
    return None