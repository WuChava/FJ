# -*- coding: UTF-8 -*-

from openpyxl import load_workbook
import csv
import datetime
import re
import os.path
import glob
import Library
from shutil import copyfile
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl import formatting, styles
import configparser

def MainPerformance():
    try:
        print('************ BEGIN ************')
        print()
        print('Step 1: Processing Template and Raw Data')
        try:
            try:
                config = configparser.ConfigParser()
                config.read('Settings.ini')
                rawdata_path=config.get('source', 'rawdata')
                if rawdata_path=="": rawdata_path=None
            except: rawdata_path=None
            
            #1.import template
            wb_file_name = 'Report\AG_Performance_Template.xlsx'
            wb_file = Library.getXlsxFile(wb_file_name, [])
            print('  -Loading file: %s ' % wb_file, end="")
            #wb = load_workbook(filename = wb_file_name, data_only=True)
            wb = load_workbook(filename = wb_file)
            for checksheet in wb.sheetnames:
                if checksheet!="Performance":
                    wb.remove(wb[checksheet])
            wb_sheet = wb["Performance"]
            ReportDate = wb_sheet.cell(column=1, row=1).value
            ReportDateLast = ReportDate - datetime.timedelta(days=1)
            ReportDateSimpleArr = str(ReportDate).split(' ',1)[0].split('-',2)
            ReportDateSimple = ReportDateSimpleArr[1]+ReportDateSimpleArr[2]
            ReportDateStr = ReportDateSimpleArr[0]+"/"+ReportDateSimpleArr[1]+"/"+ReportDateSimpleArr[2]
            ReportDateLastSimpleArr = str(ReportDateLast).split(' ',1)[0].split('-',2)
            ReportDateLastSimple = ReportDateLastSimpleArr[1]+ReportDateLastSimpleArr[2]
            #wb_sheet = wb[wb.sheetnames[0]]
            print(' => Completed')
            
            #2.import data source
            print('  -Loading file:', end="")
            
            table_summary_file=None
            if rawdata_path!=None:
                try:
                    table_summary_file_name = rawdata_path + '\**\*客服人員群組總結*.xls'
                    table_summary_file = Library.getCsvFile(table_summary_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_summary_file=None
            if table_summary_file==None:
                table_summary_file_name = 'Report\RAWDATA\**\*客服人員群組總結*.xls'
                table_summary_file = Library.getCsvFile(table_summary_file_name, [[1,1, ReportDateStr]])

            if table_summary_file!=None:
                table_summary = csv.reader(open(table_summary_file, 'r'), delimiter='\t')            
                #table_summary = getCsvFile('前日匯報-客服人員群組總結*.xls', ReportDateStr)
                table_summary_list = list(table_summary) if table_summary!=None else None
                table_summary_rows = sum(1 for row in table_summary) if table_summary_list!=None else 0
                print(' %s => Completed' % table_summary_file)
            else:
                print(' Finding "%s\" => Failed' % table_summary_file_name)

            
            #2.import data source
            print('  -Loading file:', end="")
            #table_loginout = csv.reader(open('Report\RAWDATA\前日匯報-客服人員登出登入.xls', 'r'), delimiter='\t')
            table_loginout_file=None
            if rawdata_path!=None:
                try:
                    table_loginout_file_name = rawdata_path + '\**\*客服人員登出登入*.xls'
                    table_loginout_file = Library.getCsvFile(table_loginout_file_name, [[1,1, ReportDateStr]])
                except Exception as e: table_loginout_file=None
            if table_loginout_file==None:
                table_loginout_file_name = 'Report\RAWDATA\**\*客服人員登出登入*.xls'
                table_loginout_file = Library.getCsvFile(table_loginout_file_name, [[1,1, ReportDateStr]])

            if table_loginout_file!=None:            
                table_loginout = csv.reader(open(table_loginout_file, 'r'), delimiter='\t')
                table_loginout_list = list(table_loginout) if table_loginout!=None else None
                table_loginout_rows = sum(1 for row in table_loginout) if table_loginout_list!=None else 0
                print(' %s => Completed' % table_loginout_file)
            else:
                #print(' %s => Failed, cannot found date including report date' % table_loginout_file_name)
                print(' Finding "%s\" => Failed' % table_loginout_file_name)

            
            #2.import data source
            print('  -Loading file:', end="")
            table_mail_file_name = 'Report\RAWDATA\**\*MAIL*.xlsx'
            table_mail_file = Library.getXlsxFile(table_mail_file_name, [['Str', 2, 'Closed'], ['DateStr', 4, ReportDateStr]])
            #table_mail = load_workbook(filename = 'Report\RAWDATA\MAIL.xlsx')
            if table_mail_file!=None:
                table_mail = load_workbook(filename = table_mail_file)
                table_mail_sheet = table_mail[table_mail.sheetnames[0]]
                print(' %s => Completed' % table_mail_file)
            else:
                print(' Finding "%s\" => Failed' % table_mail_file_name)

            
            #2.import data source
            print('  -Loading file:', end="")
            table_cts_file_name = '*Report\RAWDATA\**\*CTS*.xlsx'
            table_cts_file = Library.getXlsxFile(table_cts_file_name, [['DateStr', 3, ReportDateStr]])
            #table_cts_file = Library.getCtsFileName(table_cts_file_name, ReportDateStr)
            if table_cts_file!=None:
                table_cts = load_workbook(filename = table_cts_file)
                table_cts_sheet = table_cts[table_cts.sheetnames[0]]
                print(' %s => Completed' % table_cts_file)
            else:
                print(' Finding "%s\" => Failed' % table_cts_file_name)

            #3.Processing Report
            print()
            print('Step 2: Processing report calculation')

            TotalACD = 0
            TotalACW = 0
            TotalLogin = 0

            if wb_sheet.max_row>1:
                wb_sheet.cell(column=4, row=3).value = 0
                for rows in range(1, wb_sheet.max_row+1):
                    LoginID = wb_sheet.cell(column=4, row=rows).value
                    LoginName = wb_sheet.cell(column=2, row=rows).value                
                    
                    if table_summary_file!=None:                    
                        
                        mySummaryRow = Library.getRow(table_summary_list, LoginID)
                                            
                        #Process Summary Table
                        if mySummaryRow!=None:

                            print('  -Loading LoginID = %s'  % LoginID, end="")

                            #print(mySummaryRow)
                            #Start ACD Process...
                            TotalACD += int(mySummaryRow[10])
                            wb_sheet.cell(column=13, row=rows).value = datetime.timedelta(seconds=int(mySummaryRow[10]))   #ACD
                            wb_sheet.cell(column=13, row=3).value = datetime.timedelta(seconds=TotalACD)   #Total ACD
                            #End ACD Process...
                            #Start ACW Process...
                            TotalACW += int(mySummaryRow[11])
                            wb_sheet.cell(column=14, row=rows).value = datetime.timedelta(seconds=int(mySummaryRow[11]))   #ACW
                            wb_sheet.cell(column=14, row=3).value = datetime.timedelta(seconds=TotalACW)   #Total ACW
                            #End ACW Process...    
                            #Start TotalLogin Process...
                            TotalLogin += int(mySummaryRow[16])
                            wb_sheet.cell(column=12, row=rows).value = datetime.timedelta(seconds=int(mySummaryRow[16]))   #TotalLogin
                            wb_sheet.cell(column=12, row=3).value = datetime.timedelta(seconds=TotalLogin)   #Total TotalLogin
                            #End TotalLogin Process...
                            #Start ACH Process...
                            wb_sheet.cell(column=28, row=rows).value = datetime.timedelta(seconds=int(float(mySummaryRow[2])+float(mySummaryRow[3])))   #ACH
                            #End TotalLogin Process...
                                                                                        
                    
                    #Process Loginout Table
                    if table_loginout_file!=None:
                        myLoginoutRow = Library.getRow(table_loginout_list, LoginID)
                        if myLoginoutRow!=None:
                            #print(myLoginoutRow)
                            LoginSec = Library.getSec(myLoginoutRow[3])
                            LogoutSec = Library.getSec(myLoginoutRow[5])
                            wb_sheet.cell(column=10, row=rows).value = str(datetime.timedelta(seconds=LoginSec))[-5:]   #Login
                            wb_sheet.cell(column=11, row=rows).value = str(datetime.timedelta(seconds=LogoutSec))[-5:]   #Logout
                            try:
                                wb_sheet.cell(column=5, row=rows).value = int((LogoutSec-LoginSec)/60)
                            except:
                                wb_sheet.cell(column=5, row=rows).value = ""
                            wb_sheet.cell(column=20, row=rows).value = wb_sheet.cell(column=5, row=rows).value
                            #wb_sheet.cell(column=10, row=rows).value = datetime.timedelta(seconds=))   #Login
                            #wb_sheet.cell(column=11, row=rows).value = datetime.timedelta(seconds=))   #Logout
                        else:
                            myRole = wb_sheet.cell(column=1, row=rows).value
                            if myRole=='AG' or myRole=='SA':
                                wb_sheet.cell(column=5, row=rows).value = "休"
                                wb_sheet.cell(column=20, row=rows).value = "休"

                    #Process Mail Table
                    TotalPaperCounter = 0
                    if rows>3 and table_mail_file!=None:
                        myMailCounter = Library.getMailCount(table_mail_sheet, LoginName, ReportDateStr)                    
                        if myMailCounter!=None and myMailCounter>0:
                            wb_sheet.cell(column=17, row=rows).value = myMailCounter
                            TotalPaperCounter += myMailCounter

                    #Process Cts Table
                    if rows>3 and table_cts_file!=None:
                        myFacebookCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, 'Facebook')
                        if myFacebookCounter!=None and myFacebookCounter>0:
                            wb_sheet.cell(column=18, row=rows).value = myFacebookCounter
                            TotalPaperCounter += myFacebookCounter
                        
                        myTelCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, '電話')
                        if myTelCounter!=None and myTelCounter>0:
                            wb_sheet.cell(column=15, row=rows).value = myTelCounter
                            TotalPaperCounter += myTelCounter

                        myOutboundCounter = Library.getCtsCount(table_cts_sheet, LoginName, ReportDateStr, 'Outbound')
                        if myOutboundCounter!=None and myOutboundCounter>0:
                            wb_sheet.cell(column=16, row=rows).value = myOutboundCounter 
                            TotalPaperCounter += myOutboundCounter

                    if wb_sheet.cell(column=5, row=rows).value == "休" and TotalPaperCounter>0:
                        wb_sheet.cell(column=5, row=rows).value = ""
                        wb_sheet.cell(column=20, row=rows).value = ""
                        
                    if str(wb_sheet.cell(column=5, row=rows).value).isdigit():
                            wb_sheet.cell(column=4, row=3).value += 1           

                    if table_summary_file!=None and mySummaryRow!=None:
                        print(' => Completed')

                    

            #改數值格式後不用計算
            #if TotalLogin>0 and TotalACD>0:
            #    wb_sheet.cell(column=21, row=3).value = TotalACD/TotalLogin   # Total TotalACD/TotalLogin
            #if TotalLogin>0 and TotalACW>0:
            #    wb_sheet.cell(column=23, row=3).value = (TotalACD+TotalACW)/TotalLogin   # Total TotalACD/TotalLogin
            
            print()
            print('Step 3: Generating Report')
            #Generate        
            wb_sheet.title = ReportDateSimple
            Performance_FilePathName = "Report\OPPO_Agent_Performance%s.xlsx" % ReportDateSimple
            Performance_FilePathNameLast = "Report\OPPO_Agent_Performance%s.xlsx" % ReportDateLastSimple

            print('  -Creating Report to the %s'  % Performance_FilePathName, end="")

            gray_font = styles.Font(color='00A0A0A0')
            if not os.path.isfile(Performance_FilePathName) and os.path.isfile(Performance_FilePathNameLast) and ReportDateSimple[-2:]!='01':
                copyfile(Performance_FilePathNameLast, Performance_FilePathName)

            if os.path.isfile(Performance_FilePathName):
                wb_Copy = load_workbook(filename = Performance_FilePathName)
                if ReportDateSimple in wb_Copy.sheetnames:
                    wb_Copy_Sheet = wb_Copy[ReportDateSimple]    
                else:        
                    wb_Copy_Sheet = wb_Copy.copy_worksheet(wb_Copy[wb_Copy.sheetnames[0]])
                    wb_Copy_Sheet.title = ReportDateSimple
                wb_Copy_Sheet = Library.copyWorksheet(wb_sheet, wb_Copy_Sheet)
                
                wb_Copy_Sheet.conditional_formatting.add('A1:AB100', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=gray_font))            
                wb_Copy.active = len(wb_Copy.sheetnames)-1
                wb_Copy.save(Performance_FilePathName)
                wb_Copy.close()

            else:
                wb_sheet.conditional_formatting.add('A1:AB100', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=gray_font))
                wb.save(Performance_FilePathName)

            wb.close()
            print(' => Completed')    

            #print(str(datetime.timedelta(seconds=134)))
            #print("12305"[-2:])
            #print('0'.isdigit())
            #test='Yvonne Peng彭伊楺'
            #print(re.sub(r'[\x00-\x7f]',r' ',test).strip())

            #print(ReportDateLastSimple)
            #print(ReportDateStr)
        #except Exception as e:
        #    print(e)
        except:
            print('  -Loading file: => Failed ')
    #except Exception as e:
    #    print(e)
    except:
       print("Error!! Close all excel files and try again.")

    finally:
        print()
        print('************ END ************')